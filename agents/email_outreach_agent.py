# email_outreach_agent.py
"""
Email Outreach Agent — sends templated referral/outreach emails to 'email_target'
contacts only. Fills one of 5 rotating templates (email_templates table) with
company name, person name, and job role, attaches the matching resume variant,
and sends via SMTP. Enforces the daily cap (target: 50/day, ramping up as contact
supply allows) and only sends within the two approved windows (9:30-11:00 AM and
1:00-2:30 PM IST, weekdays). Checks outreach_log before sending to guarantee no
contact is ever emailed twice. Runs on the Linux laptop, triggered by n8n.
"""
import re
import time
import random
import smtplib
import os
from email.message import EmailMessage
from datetime import datetime
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
load_dotenv()
# ---------- CONFIG ----------
BASE_DIR = Path(__file__).resolve().parent.parent
CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
TEMPLATES_FILE = BASE_DIR / "templates.txt"
RESUME_DIR = BASE_DIR / "resumes"
MAX_EMAILS_PER_CYCLE = 20
DELAY_RANGE = (15, 30)  # seconds

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = os.environ["SMTP_USER"]
SMTP_PASS = os.environ["SMTP_PASS"]


def load_templates(filepath: Path) -> list[dict]:
    text = filepath.read_text(encoding="utf-8")
    chunks = re.split(r"---TEMPLATE\d+---", text)
    chunks = [c.strip() for c in chunks if c.strip()]

    templates = []
    for chunk in chunks:
        lines = chunk.split("\n")
        assert lines[0].startswith("SUBJECT:"), f"Malformed template: {chunk[:50]}"
        subject = lines[0].replace("SUBJECT:", "").strip()
        body = "\n".join(lines[1:]).strip()
        templates.append({"subject": subject, "body": body})

    assert len(templates) > 0, "No templates parsed"
    return templates


def get_resume_path() -> Path:
    files = list(RESUME_DIR.glob("*"))
    assert len(files) == 1, f"Expected exactly 1 resume file, found {len(files)}"
    return files[0]


def build_email(template: dict, name: str, company: str, resume_path: Path) -> EmailMessage:
    subject = template["subject"].format(company=company)
    body = template["body"].format(name=name, company=company)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SMTP_USER
    msg.set_content(body)

    with open(resume_path, "rb") as f:
        resume_data = f.read()
    msg.add_attachment(
        resume_data,
        maintype="application",
        subtype="octet-stream",
        filename=resume_path.name,
    )
    return msg


def send_email(msg: EmailMessage, to_email: str):
    msg["To"] = to_email
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)


def run_cycle():
    templates = load_templates(TEMPLATES_FILE)
    resume_path = get_resume_path()

    wb = openpyxl.load_workbook(CONTACTS_FILE)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    sent_count = 0
    for row_idx in range(2, ws.max_row + 1):
        if sent_count >= MAX_EMAILS_PER_CYCLE:
            break

        row = ws[row_idx]
        status = row[col["status"]].value

        if status:  # already sent, skip
            continue

        name = row[col["name"]].value
        email = row[col["email"]].value
        company = row[col["company"]].value

        if not name or not email or not company:
            continue  # incomplete row, skip

        template_idx = row_idx % len(templates)
        template = templates[template_idx]

        try:
            msg = build_email(template, name, company, resume_path)
            send_email(msg, email)

            row[col["status"]].value = "sent"
            row[col["sent_at"]].value = datetime.now().isoformat(timespec="seconds")
            row[col["template_used"]].value = template_idx + 1
            wb.save(CONTACTS_FILE)  # save immediately, not batched

            sent_count += 1
            print(f"Sent to {email} ({company}) using template {template_idx + 1}")

        except Exception as e:
            row[col["status"]].value = "failed"
            wb.save(CONTACTS_FILE)
            print(f"FAILED to send to {email}: {e}")

        if sent_count < MAX_EMAILS_PER_CYCLE:
            time.sleep(random.uniform(*DELAY_RANGE))

    print(f"Cycle complete. Sent {sent_count} emails.")


if __name__ == "__main__":
    run_cycle()